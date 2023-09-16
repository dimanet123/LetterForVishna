# -*- coding: utf-8 -*-
"""
Created on Sat Sep 16 13:45:58 2023
This programm serves a purpose of making an excel file,
which gets all marks from file ListOfMarks.xlsx,
but only if student is target_abitur. Checking is made with file ListOfStudents.xlsx.
Proramm generates file Marks.xlsx.

@author: Dimas
"""
import pandas as pnd
import numpy as np
import openpyxl
from openpyxl.styles import (
                        Alignment
                        )
from openpyxl.utils.dataframe import dataframe_to_rows


aligment_day = Alignment(horizontal='right',
                    vertical='center',
                    text_rotation = 90,
                    wrap_text = True,
                    shrink_to_fit = False,
                    indent = 0)

table = openpyxl.Workbook()


MainTable = openpyxl.load_workbook("ListOfMarks.xlsx")
tableCheck = pnd.read_excel('ListOfStudents.xlsx', sheet_name="Worksheet")
tableCheck = tableCheck[tableCheck['Целевое обучение']=="Да - ЦДП"]
table_academ = tableCheck[tableCheck['Состояние обучения']=="В академическом отпуске"]

def output_of_group(sheet_name):
    """

    Parameters
    ----------
    sheet_name : string
        This string contains a name of sheet

    Returns
    -------
    A new excel sheet
        A new sheet, which contains marks of target_students from file ListOfMarks.xlsx

    """
    width = 40
    output_of_group = pnd.read_excel('ListOfMarks.xlsx', sheet_name=sheet_name)
    output_of_group = output_of_group[output_of_group['ФИО'].isin(tableCheck['ФИО'])]
    sheet = table.create_sheet(sheet_name)
    empty_column = pnd.Series([np.nan] * len(output_of_group), name='ФИО')
    output_of_group = output_of_group.drop('№', axis=1)
    output_of_group.insert(0, 'Группа',empty_column)
    output_of_group['Группа'] = sheet_name
    def add_4_if_exists(name):
        if name in table_academ['ФИО'].values:
            return f"{name} а.о."
        return name
    output_of_group['ФИО'] = output_of_group['ФИО'].apply(lambda x: add_4_if_exists(x))
    for row in dataframe_to_rows(output_of_group, index=False, header=True):
        sheet.append(row)
    for i in range (20):
        sheet.cell(row = 1, column = 4+i).alignment = aligment_day
    sheet.column_dimensions['B'].width = width
    sheet.column_dimensions['C'].width = 20

for sheet_name in MainTable.sheetnames:
    output_of_group(sheet_name)
table.save('Marks.xlsx')
